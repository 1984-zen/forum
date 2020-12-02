from django.shortcuts import render
from exams.models import Exams, Questions, Options, Option_Users, Exam_files, Exam_Users
from accounts.models import Users
from django.http import HttpResponseRedirect
from django.urls import reverse
import datetime
import os
from django.db.utils import DatabaseError, IntegrityError
from django.contrib import messages
# from pprint import pprint
from django.http import JsonResponse
from django.http import HttpResponse
import json
from django.conf import settings
from django.utils import timezone, dateformat
from django.template.response import TemplateResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import Workbook
import re
from django.core.exceptions import ObjectDoesNotExist
import ast

def show_exam_list(request):
    user_id = request.session.get('user_id')
    exam_list = Exams.objects.all().order_by('-created_at')
    exam_ids_list = list(Exams.objects.all().values_list('id', flat = True))
    get_exam_users_id_list = Exam_Users.objects.filter(user_id = user_id).filter(count = 0).values_list('exam_id', 'id') #從user那些尚未完成的考卷中拿到Exam_Users.exam_id及Exam_Users.id
    last_answer_list = [] #紀錄user上次做到哪一題的結果 [{'exam_id': 222, 'next_answer_page': 1}]
    user_incomplete_exam_id_list = [] #紀錄user在Exam_Users表裡面尚未完成的exam_id，例如 [217, 221, 222]
    for (exam_id, exam_users_id) in get_exam_users_id_list:
        get_questions_id = Questions.objects.filter(exam_id = exam_id).values_list('id', flat=True) #拿到該exam_id中的所有Questions題目id
        get_user_answered_question_id = Option_Users.objects.filter(exam_users_id = exam_users_id).values_list('question_id', flat=True) #拿到所有user在該exam_id中回答過的question_id
        if(set(get_questions_id) == set(get_user_answered_question_id)): #如果user都填完所有答案，但尚未交卷
            last_answer_id = max(set(get_questions_id)) #就取出最後一題的question id
        else: #如果user還有1題以上的題目尚未答題
            last_answer_id = min(set(get_questions_id) - set(get_user_answered_question_id)) #尚未完成題目的最小question_id = 取最小值(所有題目的id - 使用者回答過的題目)
        page_num = list(get_questions_id).index(last_answer_id) #拿到尚未答題的頁數
        last_answer_list.append({"exam_id": exam_id, "next_answer_page": page_num + 1})
        user_incomplete_exam_id_list.append(exam_id)
    return TemplateResponse(request, 'exam_list.html', {'exam_list': exam_list, 'last_answer_list': last_answer_list, 'user_incomplete_exam_id_list': user_incomplete_exam_id_list, 'exam_ids_list': exam_ids_list})

def show_exam_user_list(request, exam_id):
    user_list = Exam_Users.objects.filter(exam_id = exam_id).exclude(count = 0).select_related('user').values('user__id', 'user__name', 'date', 'count').order_by('user__name', '-created_at')
    exam = Exams.objects.get(id = exam_id)
    return TemplateResponse(request, 'exam_user_list.html', {'user_list': user_list, 'exam': exam})

def new_exam(request):
    return TemplateResponse(request, 'new_exam.html', {}) #這裡是空object是為了能接到middleware的process_template_response()的response

def update_question(request, exam_id, question_id):
    if(request.method == "POST"): #如果user提交更新某個問題,該題選項,跳至某一題的話
        question = request.POST.get("question") #前端user修正question text
        update_question = Questions.objects.filter(id = question_id) #更新question
        update_question.update(question = question) #存入更新

        options_id_list = Options.objects.filter(question_id = question_id).values_list('id', flat = True) #拿到該question下的所有option的id
        option_list = request.POST.getlist("option[]") #前端user修正的所有option text
        is_answer_list = request.POST.getlist("is_answer[]")  #前端user修正的所有is_answer
        for i in range(len(options_id_list)): #for迴圈去一個個判斷option的checkbox是否被user打過勾，打過勾就會把option_id紀錄在is_answer_list裡面，然後再去對照如果option_id也出現在is_answer_list就會顯示True
            has_option_id = str(options_id_list[i]) in is_answer_list # 回傳True or False
            update_is_answer = Options.objects.filter(id = options_id_list[i])
            update_is_answer.update(is_answer = has_option_id, option = option_list[i])
        if('next' in request.GET): #如果前端的<form action>裡面有個?next={{request.GET.page}}這個page就是記住從前一頁pass過來的頁數，注意這個key是str
            return HttpResponseRedirect(reverse("add_more_questions", kwargs={"exam_id": exam_id}) + "?page=" + request.GET.get("next")) #next就是頁數值，就會redirect回到?page=上一個畫面的頁數
        else: #如果<form action>沒有?next這個key的話
            return HttpResponseRedirect(reverse("add_more_questions", kwargs={"exam_id": exam_id}))
    exam = Exams.objects.get(id = exam_id) #if(request.method == "GET")顯示update_question.html畫面
    questions = exam.questions.all()
    question = questions.get(id = question_id)
    next_question_id_list = Options.objects.filter(question_id = question_id).values_list('id', 'next_question_id')
    next_question_list = []
    for (option_id, next_question_id) in next_question_id_list:
        try:
            next_question = Questions.objects.get(id = next_question_id)
            next_question_list.append({"next_question": next_question, "option_id": option_id})
        except ObjectDoesNotExist:
            next_question_list.append({"next_question": None, "option_id": option_id})
    return TemplateResponse(request, 'update_question.html', {'exam': exam, 'questions': questions, 'question': question, 'next_question_list': next_question_list}) #這裡是空object是為了能接到middleware的process_template_response()的response

def update_next_question_id(request, exam_id, question_id, option_id):
    if(request.method == "POST"):
        next_quetion_id = request.POST.get("next_question_id")
        if(next_quetion_id == "jump_to_next_default_question"): #前端要回復正常跳到下一個問題，後端判斷此值有兩種可能，一種是不是最後一個問題，就跳到下一題question id + 1, 否則給None這樣DB就會把跳到下一個問題給null值
            try:
                query_next_question_result = Questions.objects.get(id = question_id + 1).id #去query判斷是否真的有這個question_id + 1的資料，如果有就抓出這個question id
            except ObjectDoesNotExist:
                query_next_question_result = None #如果id查無此row就給None，讓DB在next_question_id寫入null
            update_option_next_question_id = Options.objects.filter(id = option_id)
            update_option_next_question_id.update(next_question_id = query_next_question_result)
            return HttpResponseRedirect(reverse("update_question", kwargs={"exam_id": exam_id, 'question_id': question_id}))
        else:
            pattern = '([0-9]+)-?' #如果是指定某個問題就用正則拿到question_id
            match = re.search(pattern, next_quetion_id)
            if(match): #如果有抓到就會有值，如果沒有就會回傳None
                ans = match.group(1) #因為我要的question id在match裡面的一個group裡面
                update_option_next_question_id = Options.objects.filter(id = option_id)
                update_option_next_question_id.update(next_question_id = ans)
                return HttpResponseRedirect(reverse("update_question", kwargs={"exam_id": exam_id, 'question_id': question_id}))
    option = Options.objects.get(id = option_id)
    exam = Exams.objects.get(id = exam_id)
    questions = exam.questions.all()
    question = questions.get(id = question_id)
    return TemplateResponse(request, 'update_option_next_question_id.html', {'exam': exam, 'option': option, 'questions': questions, 'question': question})

def get_ajax_answers_options(request):
    user_id = request.session.get('user_id')
    question = request.POST.get("question")
    option_list = request.POST.get("option_list").split(',') #因為從前端接收到的值= True,False所以用split(',')變成['True', 'False']
    answer_list = request.POST.get("answer_list").split(',') #因為從前端接收到的值= True,False所以用split(',')變成['True', 'False']
    if(request.POST.get("exam_id")): #如果已經創立好exam只是要增加問題跟選項
        #開始新增問題
        exam_id = request.POST.get("exam_id")
        create_question = Questions(question = question, exam_id = exam_id)
        create_question.save()
        #開始要幫上一個question更新next_question_id
        last_question_id = Questions.objects.filter(exam_id = exam_id).order_by('-id')[1].id #找到上一個question_id
        last_options = Options.objects.filter(question_id = last_question_id).filter(next_question_id = None) #找到它底下的options是屬於沒有指定任何next_question_id
        update_last_options = last_options.update(next_question_id = create_question.id) #就幫他們都指定剛剛創建好的question_id

        for i in range(len(option_list)):
            create_option = Options(option = option_list[i], is_answer = answer_list[i], question_id = create_question.id) #一開始next_question_id是null
            create_option.save()

        if(request.FILES.getlist('media_file')):
            files = request.FILES.getlist('media_file')
            for file in files:
                fname, file_relative_path = handle_uploaded_media_file(file)
                create_file_path = Exam_files(name = fname, file_path = file_relative_path, type = "media", exam_id = exam_id, question_id = create_question.id)
                create_file_path.save()
        
        if(request.FILES.getlist('image_list')):
            images = request.FILES.getlist('image_list')
            for image in images:
                fname, file_relative_path = handle_uploaded_image_file(image)
                create_file_path = Exam_files(name = fname, file_path = file_relative_path, type = "image", exam_id = exam_id, question_id = create_question.id)
                create_file_path.save()
        return JsonResponse({'exam_id': exam_id}) #回傳data的值給前端的Ajax，並就停留在原url
    else: #如果是新建立exam
        exam_title = request.POST.get("exam_title")
        create_exam = Exams(user_id = user_id, name = exam_title)
        create_exam.save()
        create_question = Questions(question = question, exam_id = create_exam.id)
        create_question.save()
        
        for i in range(len(option_list)):
            create_option = Options(option = option_list[i], is_answer = answer_list[i], question_id = create_question.id) #一開始next_question_id是null
            create_option.save()
        
        if(request.FILES.getlist('media_file')):
            files = request.FILES.getlist('media_file')
            for file in files:
                fname, file_relative_path = handle_uploaded_media_file(file)
                create_file_path = Exam_files(name = fname, file_path = file_relative_path, type = "media", exam_id = create_exam.id, question_id = create_question.id)#拿取剛建立完的create_exam.id
                create_file_path.save()

        if(request.FILES.getlist('image_list')):
            images = request.FILES.getlist('image_list')
            for image in images:
                fname, file_relative_path = handle_uploaded_image_file(image)
                create_file_path = Exam_files(name = fname, file_path = file_relative_path, type = "image", exam_id = create_exam.id, question_id = create_question.id)#拿取剛建立完的create_exam.id
                create_file_path.save()
        return JsonResponse({'exam_id': create_exam.id})

def handle_uploaded_media_file(f):
    timestamp = str(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
    file_relative_path = timestamp + '_' + f.name
    file_path = os.path.join(os.path.dirname(__file__),'upload_file/media', file_relative_path)
    with open(file_path, 'wb') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return f.name, file_relative_path

def handle_uploaded_image_file(i):
    timestamp = str(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))
    file_relative_path = timestamp + '_' + i.name
    file_path = os.path.join(os.path.dirname(__file__),'upload_file/media', file_relative_path)
    with open(file_path, 'wb') as destination:
        for chunk in i.chunks():
            destination.write(chunk)
    return i.name, file_relative_path

def add_exam_questions(request, exam_id):
    exam = Exams.objects.get(id = exam_id)
    questions = exam.questions.all().order_by('created_at')
    page = request.GET.get('page', 1)
    paginator = Paginator(questions, 5)
    questions_in_page = paginator.page(page)
    videos = exam.exam_files.filter(type = 'media')
    images = exam.exam_files.filter(type = 'image')
    return TemplateResponse(request, 'add_exam_questions.html', {'exam_id': exam_id, 'exam': exam, 'questions': questions, 'videos': videos, 'images': images, 'questions_in_page': questions_in_page})

def show_exam(request, exam_id):
    user_id = request.session.get('user_id')
    exam = Exams.objects.get(id = exam_id)
    questions = exam.questions.all().order_by('created_at')
    page = request.GET.get('page', 1)
    paginator = Paginator(questions, 1)
    questions_in_page = paginator.page(page)
    videos = exam.exam_files.filter(type = 'media')
    images = exam.exam_files.filter(type = 'image')
    has_user_incomplete_record = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 0).count() #找到user未完成的考卷count = 0
    if(has_user_incomplete_record): #如果user正在做尚未完成的考卷
        #以下list用來前端判斷user有無回答過question或option，如果有前端就顯示出來
        exam_users_id = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 0)[0].id #查詢是哪個exam_users的id
        user_has_been_answered_question_list = list(Option_Users.objects.filter(exam_users_id = exam_users_id).values_list('question_id', flat=True)) #把所有關於這張位完成考卷所有user回答過的question_id都列出
        user_has_been_answered_option_list = list(Option_Users.objects.filter(exam_users_id = exam_users_id).values_list('option_id', flat=True)) #把所有關於這張位完成考卷所有user回答過的option_id都列出
    else: #如果沒有就回傳空list給前端不顯示
        user_has_been_answered_question_list = [] #user沒有回答過任何question
        user_has_been_answered_option_list = [] #user沒有回答過任何option
    return TemplateResponse(request, 'show_exam.html', {'user_id': user_id, 'exam_id': exam_id, 'exam': exam, 'exam_id': exam_id, 'questions_in_page': questions_in_page, 'videos': videos, 'images': images, 'user_has_been_answered_question_list': user_has_been_answered_question_list, 'user_has_been_answered_option_list': user_has_been_answered_option_list})

def user_answers(request, exam_id):
    user_id = request.session.get('user_id')
    exam_id = exam_id
    questions_id_list = list(Questions.objects.filter(exam_id = exam_id).values_list('id', flat = True)) #拿到所有question的id
    current_page = request.POST.get('current_page') #從前端的input有個name叫做"current_page"，印出型別是str，所以next_page要加1之前要轉int()
    user_answers_list = request.POST.getlist("user_answers[]")
    if(not user_answers_list): #如果考生沒有勾選答案就提交就回到原本題目
        messages.error(request, "You have not selected any answer!")
        return HttpResponseRedirect(reverse('show_exam', kwargs={"exam_id": exam_id}) + "?page=" + current_page) #跳回到同一題
    
    has_incomplete_exam_record = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 0).count()#找出user在交卷紀錄裡面是否有count = 0的紀錄
    if(has_incomplete_exam_record):#如果user有尚未交卷的紀錄就讓user繼續作答
        for user_answer_str in user_answers_list: #user_answer = {'option_id': 110762, 'next_question_id': 15543, 'question_id': 15542}
            user_answer_dict = ast.literal_eval(user_answer_str) #ast.literal_eval是拿來str轉換dict
            #開始判斷user送出答案要轉跳的頁面是否為最後一題
            if(user_answer_dict['next_question_id']): #如果不是最後一題，判斷要去的頁面是第幾頁
                next_question_page = questions_id_list.index(user_answer_dict['next_question_id']) #找到next_question_id是在全部questions的第幾個位置，next_question_id_index就是要轉跳到第幾頁
            else: #如果是最後一題則停留在原畫面
                next_question_page = int(current_page) - 1 #因為最下面的return有 + 1, 所以這裡預先 - 1
            exam_users_id = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 0)[0].id #目前user尚未完成的考卷id
            create_option_users = Option_Users(user_id = user_id, option_id = user_answer_dict['option_id'], question_id = user_answer_dict['question_id'], exam_id = exam_id, exam_users_id = exam_users_id) #把user答案寫入DB
            create_option_users.save()
        messages.success(request, "Answer has been sent successfully!")
    else:#表示user之前都交卷了，給他創建新的考卷，count = 0
        create_exam_users = Exam_Users(user_id = user_id, exam_id = exam_id, date = datetime.datetime.now(), count = 0)
        create_exam_users.save() #user作答完就紀錄在exam_user上，用來計算他是第幾次應考
        for user_answer_str in user_answers_list:
            user_answer_dict = ast.literal_eval(user_answer_str) #ast.literal_eval是拿來str轉換dict
            #開始判斷user送出答案要轉跳的頁面是否為最後一題
            if(user_answer_dict['next_question_id']): #如果不是最後一題，判斷要去的頁面是第幾頁
                next_question_page = questions_id_list.index(user_answer_dict['next_question_id']) #找到next_question_id是在全部questions的第幾個位置，next_question_id_index就是要轉跳到第幾頁
            else: #如果是最後一題則停留在原畫面
                next_question_page = int(current_page) - 1 #因為最下面的return有 + 1, 所以這裡預先 - 1
            exam_users_id = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 0)[0].id #目前user尚未完成的考卷id
            create_option_users = Option_Users(user_id = user_id, option_id = user_answer_dict['option_id'], question_id = user_answer_dict['question_id'], exam_id = exam_id, exam_users_id = exam_users_id) #把user答案寫入DB
            create_option_users.save()
        messages.success(request, "Answer has been sent successfully!")
    return HttpResponseRedirect(reverse('show_exam', kwargs={"exam_id": exam_id}) + "?page=" + str(int(next_question_page) + 1)) #跳到下一題，因為index是從0開始所以要 + 1

def user_exam_completed(request, exam_id, user_id):
    user_id = request.session.get('user_id')
    exam_id = exam_id
    user_has_completed_exam_count = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 1).count() #算出已完成考試的數量
    update_exam_users_count_to_1 = Exam_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(count = 0) #找到user未完成的考卷count = 0
    update_exam_users_count_to_1.update(count = user_has_completed_exam_count + 1) #已完成考試的MAX數量+1
    return HttpResponse("Finished!! All answers have been sent successfully.<a href=\"/exams\">Go back to exam list</a>")

def show_user_exam_result(request, exam_id, user_id, user_exam_count):
    exam = Exams.objects.get(id = exam_id)
    questions = Questions.objects.filter(exam_id = exam_id)
    videos = exam.exam_files.filter(type = 'media')
    images = exam.exam_files.filter(type = 'image')
    page = request.GET.get('page', 1)
    paginator = Paginator(questions, 1)
    questions_in_page = paginator.page(page)

    exam_users_id = Exam_Users.objects.filter(user_id = user_id).filter(exam_id =exam_id).filter(count = user_exam_count)[0].id
    user_answers = Option_Users.objects.filter(user_id = user_id).filter(exam_id = exam_id).filter(exam_users_id = exam_users_id)
    user_answer_list = [] #user再這張exam所勾選過的選項
    for user in user_answers:
        user_answer_list.append(user.option_id) #把所有user對這張exam的所有勾選的option都存進到user_answer_list
    return TemplateResponse(request, 'show_user_exam_result.html', {'questions_in_page': questions_in_page, 'user_answer_list': user_answer_list, 'videos': videos, 'images': images, 'exam': exam})

def delete_question(request, exam_id, question_id):
    question = Questions.objects.filter(id = question_id)
    question.delete()
    return HttpResponseRedirect(reverse("add_more_questions", kwargs={"exam_id": exam_id}))

def delete_question_media(request, exam_id, question_id, question_media_id):
    question_media_file = Exam_files.objects.get(id = question_media_id)
    question_media_file.delete()
    return HttpResponseRedirect(reverse("add_more_questions", kwargs={"exam_id": exam_id}))

def delete_question_image(request, exam_id, question_id, question_image_id):
    question_image_file = Exam_files.objects.get(id = question_image_id)
    question_image_file.delete()
    return HttpResponseRedirect(reverse("add_more_questions", kwargs={"exam_id": exam_id}))

def delete_option(request, exam_id, option_id):
    option = Options.objects.get(id = option_id)
    option.delete()
    return HttpResponseRedirect(reverse("add_more_questions", kwargs={"exam_id": exam_id}))

def delete_exam(request, exam_id):
    exam = Exams.objects.get(id = exam_id)
    exam.delete()
    return HttpResponseRedirect(reverse("show_exam_list"))

def export_user_answer_xls(request, exam_id, user_id, user_exam_count):
    user_name = Users.objects.get(id = user_id).name
    exam_date = Exam_Users.objects.filter(exam_id = exam_id).filter(user_id = user_id).filter(count = user_exam_count).values('date')[0]['date'].strftime('%Y-%m-%d_%H%M')
    exam_title = Exams.objects.get(id = exam_id).name
    response = HttpResponse(content_type='application/ms-excel') #匯出excel是用response的content-type
    response['Content-Disposition'] = f"attachment; filename = {user_name}_answer_{exam_date}.xlsx" #匯出excel的檔名

    wb = Workbook()
    wb.create_sheet()
    ws = wb.active #拿到地1個sheet[0]
    ws.title = f"{user_name}_answer_{exam_date}" #命名sheet名稱

    questions = Questions.objects.filter(exam_id = 217).values('id') #考捲id=217
    #[{'id': 1724}, {'id': 1725}...]
    user_answers = Option_Users.objects.filter(exam_id = 217).filter(user_id = user_id).select_related('option').values('question_id','option__option')
    #<QuerySet [{'question_id': 1728, 'option__option': 'Apical 3-chamber–color'},...]

    #開始寫入excel
    #寫入欄位
    ws.cell(1, 1).value = '試卷名稱_Exam_title' # row=1, col=1欄位名稱'試卷名稱_Exam_title'
    ws.cell(2, 1).value = 'Question' # row=2, col=1欄位名稱'Question'
    ws.cell(2, 2).value = 'Answers' # row=2, col=2欄位名稱'Answers'

    #寫入內容
    ws.cell(1, 2).value = exam_title # row=1, col=2'Exam_title'
    for i in range(len(questions)): #印出題目的流水號從1~13809
        ws.cell(i+3, 1).value = i + 1 #question流水號從1開始
        for user_answer in user_answers: #user_answers[]
            if(user_answer['question_id'] == questions[i]['id']): #檢查user有回答這個question的話...就紀錄user勾選的選項
                count_user_answer_has_more_than_one = Option_Users.objects.filter(exam_id = exam_id).filter(user_id = user_id).filter(question_id = user_answer['question_id']).values('question_id','option__option')#如果user複選的話...
                for j in range(count_user_answer_has_more_than_one.count()): #用for迴圈去橫向紀錄在excel表格
                    ws.cell(i+3, 2+j).value = count_user_answer_has_more_than_one[j]['option__option'] #user勾選的ansers從row=3, col=2...col=2+j寫入
    wb.save(response) #存檔
    return response #匯出到瀏覽器下載